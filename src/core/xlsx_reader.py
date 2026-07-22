#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_reader.py — Read-only-Zugriff auf Excel-Tabellen (.xlsx / .xls)

Keine Qt-Abhängigkeiten; vollständig ohne GUI testbar.
Gibt bei Leseproblemen klar typisierte Fehlerobjekte zurück (kein Crash).
"""
from __future__ import annotations

import os
from dataclasses import dataclass
from typing import Any

# Optionale Imports — fehlendes Paket führt zum Fallback, nicht zum Abbruch
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    xlrd = None  # type: ignore[assignment]
    HAS_XLRD = False


# Maximale Zeilen/Spalten für die Vorschau
MAX_ROWS: int = 100
MAX_COLS: int = 50


class XlsxReadError(Exception):
    """Wird geworfen, wenn die Datei nicht gelesen werden kann."""


@dataclass
class WorkbookMeta:
    """Ergebnis des Metadaten-Lesevorgangs (Blattnamen, kein Zelleninhalt)."""
    sheets: list[str]
    active_sheet: str
    error: str | None = None


def read_workbook_meta(path: str) -> WorkbookMeta:
    """
    Liest nur die Arbeitsblatt-Namen einer Excel-Datei.
    Lädt keine Zellinhalte — sehr schnell, auch bei großen Dateien.
    Gibt bei Fehler ein WorkbookMeta mit gesetztem `error`-Feld zurück.
    """
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        return _read_xls_meta(path)

    if not HAS_OPENPYXL:
        return WorkbookMeta(sheets=[], active_sheet="", error="openpyxl nicht installiert")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheets = list(wb.sheetnames)
            active = sheets[0] if sheets else ""
            return WorkbookMeta(sheets=sheets, active_sheet=active)
        finally:
            wb.close()
    except Exception as exc:
        return WorkbookMeta(sheets=[], active_sheet="", error=str(exc))


def read_workbook_sheet(
    path: str,
    sheet_name: str,
    max_rows: int = MAX_ROWS,
    max_cols: int = MAX_COLS,
) -> list[list[Any]]:
    """
    Liest ein einzelnes Arbeitsblatt und gibt die Zellen als Zeilenliste zurück.
    Maximal `max_rows` Zeilen und `max_cols` Spalten werden geladen.
    Bei Fehler wird XlsxReadError geworfen.
    """
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        return _read_xls_sheet(path, sheet_name, max_rows, max_cols)

    if not HAS_OPENPYXL:
        raise XlsxReadError("openpyxl nicht installiert")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            rows: list[list[Any]] = []
            for row_tuple in ws.iter_rows(
                max_row=max_rows, max_col=max_cols, values_only=True
            ):
                rows.append(list(row_tuple))
                if len(rows) >= max_rows:
                    break
            return rows
        finally:
            wb.close()
    except XlsxReadError:
        raise
    except Exception as exc:
        raise XlsxReadError(f"Lesefehler: {exc}") from exc


# ── .xls-Fallback (xlrd) ────────────────────────────────────────────────────

def _read_xls_meta(path: str) -> WorkbookMeta:
    if not HAS_XLRD:
        return WorkbookMeta(
            sheets=[], active_sheet="",
            error=".xls-Vorschau benötigt xlrd (nicht installiert)"
        )
    try:
        wb = xlrd.open_workbook(path)
        sheets = wb.sheet_names()
        active = sheets[0] if sheets else ""
        return WorkbookMeta(sheets=list(sheets), active_sheet=active)
    except Exception as exc:
        return WorkbookMeta(sheets=[], active_sheet="", error=str(exc))


def _read_xls_sheet(
    path: str,
    sheet_name: str,
    max_rows: int,
    max_cols: int,
) -> list[list[Any]]:
    if not HAS_XLRD:
        raise XlsxReadError(".xls-Vorschau benötigt xlrd (nicht installiert)")
    try:
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name(sheet_name)
        rows: list[list[Any]] = []
        for r in range(min(ws.nrows, max_rows)):
            row = [ws.cell_value(r, c) for c in range(min(ws.ncols, max_cols))]
            rows.append(row)
        return rows
    except XlsxReadError:
        raise
    except Exception as exc:
        raise XlsxReadError(f"Lesefehler .xls: {exc}") from exc
